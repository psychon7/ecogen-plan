"""DocumentService — deterministic document generation.

Produces PDF, XLSX, DOCX, and ZIP evidence packs using Jinja2
templates.  No LLM dependency — all content comes from structured
data produced by the calculation and extraction steps.
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

logger = logging.getLogger(__name__)


class DocumentService:
    """Generates audit-ready LEED evidence documents."""

    def __init__(self, output_dir: str = "/tmp/leed-documents") -> None:
        self._output_dir = Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)

    # ----- XLSX generation ---------------------------------------------------

    def generate_compliance_xlsx(
        self,
        credit_code: str,
        project_id: str,
        fixture_data: list[dict[str, Any]],
        calculation_results: dict[str, Any],
        occupancy_data: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Generate a formatted Excel workbook for a credit."""
        wb = Workbook()

        # Sheet 1: Fixture / Product Compliance
        ws1 = wb.active
        ws1.title = "Compliance Table"
        headers = ["Type", "Manufacturer", "Model", "Quantity", "Baseline Rate", "Design Rate", "Status"]
        ws1.append(headers)
        for item in fixture_data:
            ws1.append([
                item.get("fixture_type", item.get("product_name", "")),
                item.get("manufacturer", ""),
                item.get("model", item.get("model_number", "")),
                item.get("quantity", ""),
                item.get("baseline_rate", ""),
                item.get("design_rate", item.get("design_flow_rate", "")),
                item.get("status", ""),
            ])

        # Sheet 2: Calculations
        ws2 = wb.create_sheet("Calculations")
        ws2.append(["Metric", "Value"])
        for k, v in calculation_results.items():
            if not isinstance(v, (list, dict)):
                ws2.append([k, str(v)])

        # Sheet 3: Occupancy (if provided)
        if occupancy_data:
            ws3 = wb.create_sheet("Occupancy")
            ws3.append(["Parameter", "Value"])
            for k, v in occupancy_data.items():
                ws3.append([k, str(v)])

        filename = f"{credit_code}_{project_id}_compliance.xlsx"
        filepath = self._output_dir / filename
        wb.save(str(filepath))

        checksum = self._file_checksum(filepath)
        logger.info("Generated XLSX: %s (checksum=%s)", filepath, checksum)

        return {
            "format": "xlsx",
            "filename": filename,
            "path": str(filepath),
            "checksum": checksum,
            "file_size_bytes": filepath.stat().st_size,
        }

    # ----- Evidence manifest -------------------------------------------------

    def generate_evidence_manifest(
        self,
        project_id: str,
        credit_code: str,
        documents: list[dict[str, Any]],
        evidence_items: list[dict[str, Any]],
        source_snapshots: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Generate the JSON evidence manifest for a credit package."""
        manifest = {
            "project_id": project_id,
            "credit_code": credit_code,
            "generated_at": datetime.utcnow().isoformat(),
            "documents": documents,
            "evidence_count": len(evidence_items),
            "source_count": len(source_snapshots),
            "evidence_items": evidence_items,
            "source_snapshots": source_snapshots,
        }

        manifest_json = json.dumps(manifest, sort_keys=True, indent=2)
        manifest["manifest_checksum"] = hashlib.sha256(manifest_json.encode()).hexdigest()

        filename = f"{credit_code}_{project_id}_manifest.json"
        filepath = self._output_dir / filename
        filepath.write_text(manifest_json)

        logger.info("Generated manifest: %s", filepath)

        return {
            "format": "json",
            "filename": filename,
            "path": str(filepath),
            "checksum": manifest["manifest_checksum"],
        }

    # ----- Helpers -----------------------------------------------------------

    @staticmethod
    def _file_checksum(path: Path) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
        return h.hexdigest()
